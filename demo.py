import random
import numpy as np
from openpyxl import load_workbook
import pandas as pd
from utils import OptimalFunc, single_quality_cal
import random
from tqdm import tqdm
seed=42
random.seed(seed)
np.random.seed(seed)

categories = ['person', 'bicycle', 'car', 'motorcycle', 'airplane', 'bus', 'train', 'truck', 'boat', 'traffic light', 'fire hydrant', 'stop sign', 'parking meter', 'bench', 'bird', 'cat', 'dog', 'horse', 'sheep', 'cow', 'elephant', 'bear', 'zebra', 'giraffe', 'backpack', 'umbrella', 'handbag', 'tie', 'suitcase', 'frisbee', 'skis', 'snowboard', 'sports ball', 'kite', 'baseball bat', 'baseball glove', 'skateboard', 'surfboard', 'tennis racket', 'bottle', 'wine glass', 'cup', 'fork', 'knife', 'spoon', 'bowl', 'banana', 'apple', 'sandwich', 'orange', 'broccoli', 'carrot', 'hot dog', 'pizza', 'donut', 'cake', 'chair', 'couch', 'potted plant', 'bed', 'dining table', 'toilet', 'tv', 'laptop', 'mouse', 'remote', 'keyboard', 'cell phone', 'microwave', 'oven', 'toaster', 'sink', 'refrigerator', 'book', 'clock', 'vase', 'scissors', 'teddy bear', 'hair drier', 'toothbrush']


class PromptBandit:
    
    def __init__(self, category_number, n_class_slots,n_prompt_slots, total_iterations, method, epsilon=None):
        self.method = method
        if self.method=="epsilon_first" and epsilon is None:
            print("The value of epsilon is empty!")
            exit()
        self.epsilon = epsilon
        self.category_number = category_number
        self.n_class_slots = n_class_slots 
        self.n_prompt_slots = n_prompt_slots 
        self.total_iterations = total_iterations
        self.quality = np.zeros(category_number)
        self.ucb_quality = np.zeros(category_number)
        self.total_revenues = np.zeros(category_number) 
        self.shows = np.zeros(category_number)

    def select_class(self, cur_iteration=None):
        if self.method == "greedy":
            chosen_class_list = np.argsort(-self.quality)[:self.n_class_slots] # select top-K sellers
        if self.method == "cucb":
            chosen_class_list = np.argsort(-self.ucb_quality)[:self.n_class_slots]
        elif self.method == "random":
            chosen_class_list = random.sample(range(0, self.category_number), self.n_class_slots)
        elif self.method == "optimal":
            chosen_class_list = [20,17,30,31]
        elif self.method == "epsilon_first":
            if cur_iteration < self.epsilon * self.total_iterations:
                chosen_class_list = random.sample(range(0, self.category_number), self.n_class_slots)
            elif cur_iteration >= self.epsilon * self.total_iterations:
                chosen_class_list = np.argsort(-self.quality)[:self.n_class_slots]
                
        return chosen_class_list

    def quality_update(self, prompt_slots, selected_class):
    
        image_qualities = [0.0 for _ in range(self.category_number)]
        chosen_TICorrelations = list() 
        categories_qualities = [0.0 for _ in range(self.category_number)]
        
        for c_idx in selected_class:
            workbook = load_workbook(filename='/newdata/PromptPricing/Exp/cls/' + "class" + str(c_idx+1) + '.xlsx', read_only=True)
            sheet = workbook.active
            
            if c_idx in range(0,20):
                begin_num = 1 + 4 * random.randint(0, sheet.max_row/4 - prompt_slots)
            elif c_idx in range(20, 40):
                begin_num = 2 + 4 * random.randint(0, sheet.max_row/4 - prompt_slots)
            elif c_idx in range(40, 60):
                begin_num = 3 + 4 * random.randint(0, sheet.max_row/4 - prompt_slots)
            elif c_idx in range(60, 80):
                begin_num = 4 + 4 * random.randint(0, sheet.max_row/4 - prompt_slots)
                
            list_ClipIQA, list_Brisque, list_ClipScore, list_quality =list(), list(), list(), list()

            for row in sheet.iter_rows(min_row=begin_num, max_row=begin_num+4*(prompt_slots-1), min_col=1, max_col=5):
                # print(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)
                if (row[0].row - begin_num) % 4 == 0:
                    list_ClipIQA.append(row[2].value)
                    list_Brisque.append(row[3].value)
                    list_ClipScore.append(row[4].value)
                    # keys = ["Model", "EP", "ClipIQA", "Brisque", "ClipScore"]
                    list_quality.append(single_quality_cal(row[2].value, row[3].value, row[4].value))
                
            CurMean_ClipIQA = sum(list_ClipIQA)*1.0 / prompt_slots
            CurMean_Brisque = sum(list_Brisque)*1.0 / prompt_slots
            CurMean_ClipScore = sum(list_ClipScore)*1.0 / prompt_slots
            CurMean_Quality = sum(list_quality)*1.0/prompt_slots
            
            if len(selected_class) == self.category_number:
                self.quality[c_idx] = CurMean_Quality
            else:
                self.quality[c_idx] = (self.quality[c_idx]*self.shows[c_idx] + sum(list_quality))*1.0 / (self.shows[c_idx] + prompt_slots)
            
            self.shows[c_idx] += prompt_slots 
            self.total_revenues[c_idx] += sum(list_quality)

            image_qualities[c_idx] = (CurMean_ClipIQA - CurMean_Brisque*1.0/200 + 2)/4  # [0,1]
            chosen_TICorrelations.append((CurMean_ClipScore*1.0/100+1)/2)  # [0,1]
            categories_qualities[c_idx] = CurMean_Quality

        for c_idx in range(self.category_number):
            self.ucb_quality[c_idx] = self.quality[c_idx] + np.sqrt(((self.n_class_slots+1) * np.log(sum(self.shows))) / self.shows[c_idx])
  
        return image_qualities, chosen_TICorrelations, categories_qualities


def COL_HSG(category_number, n_class_slots, n_prompt_slots, total_iterations, HyperData, selection_method, epsilon=None):

    ss_values, PT_values, PGT_values = [], [], []
    bandit = PromptBandit(category_number, n_class_slots, n_prompt_slots, total_iterations, selection_method, epsilon)

    image_qualities, chosen_TICorrelations, categories_qualities = bandit.quality_update(n_prompt_slots, range(category_number))

    sellers_profit_iterations, platform_profit_iterations, consumer_profit_iterations, total_revenues_iterations = [], [], [], []

    for _, cur_iteration in tqdm(enumerate(range(total_iterations))):
        # print("cur_iteration", cur_iteration)
        chosen_categories = bandit.select_class(cur_iteration)
        
        categoty_profit_list, platform_profit, consumer_profit, sigma_best, PT_best, PGT_best = OptimalFunc(chosen_categories, image_qualities, chosen_TICorrelations, categories_qualities, HyperData, category_number)

        ss_values.extend(sigma_best)
        PT_values.append(PT_best)
        PGT_values.append(PGT_best)

        sellers_profit_iterations.append(sum(categoty_profit_list))
        platform_profit_iterations.append(platform_profit)
        consumer_profit_iterations.append(consumer_profit)
        total_revenues_iterations.append(sum(bandit.total_revenues))
        
        image_qualities, chosen_TICorrelations, categories_qualities = bandit.quality_update(n_prompt_slots, chosen_categories)
    
    return bandit.shows, total_revenues_iterations, sellers_profit_iterations, platform_profit_iterations, consumer_profit_iterations


def COL_save_to_excel(method, total_revenues_iterations, sellers_profit_iterations, 
                       platform_profit_iterations, consumer_profit_iterations, category_number, n_class_slots, epsilon=None):

    method_items = []
    for cur_iter in range(total_iterations):
        
        cur_item = [cur_iter, total_revenues_iterations[cur_iter], sellers_profit_iterations[cur_iter], platform_profit_iterations[cur_iter], consumer_profit_iterations[cur_iter]]
        method_items.append(cur_item)
        
    df = pd.DataFrame(method_items, columns=['Iter', 'Revenue', 'Seller_Profit', 'Platform_Profit', 'Consumer_Profit'])
    para = "_N_" + str(category_number) + "_K_" + str(n_class_slots)
    pre_fix = method + "_" + str(epsilon) + para if epsilon != None else method + para
    df.to_excel("/newdata/PromptPricing/Exp/results_final/101_COL_" + pre_fix + ".xlsx", index=False)


if __name__ == "__main__":

    category_number = 80
    n_class_slots = 4
    n_prompt_slots = 4
    total_iterations = 5000
    
    HyperFilepath = "/newdata/PromptPricing/Exp/all_q_bar_i_t_extended.xlsx"
    HyperData = pd.read_excel(HyperFilepath)

    selection_method = "greedy"

    total_shows, total_revenues_iterations, sellers_profit_iterations, platform_profit_iterations, consumer_profit_iterations = COL_HSG(category_number, n_class_slots, n_prompt_slots, total_iterations, HyperData, selection_method)
    COL_save_to_excel(selection_method, total_revenues_iterations, sellers_profit_iterations, platform_profit_iterations, consumer_profit_iterations, category_number, n_class_slots)
